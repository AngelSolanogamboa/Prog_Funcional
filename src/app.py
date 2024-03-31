import openpyxl 
from tabulate import tabulate
from functools import reduce 
import time

# abrir excel
def excel_dataframe():
    return openpyxl.load_workbook("Inventario.xlsx")

# mostrar menu de Operaciones y devuelve el numero de operacion seleccionado 
def selectOper():
    print("\nMenu de Operaciones")
    print("1.Inventario inicial\n2.Venta\n3.Compra\n4.Mostrar Inventario")
    return int(input("Seleccione una Operacion: "))

# menu de Productos y devuelve el numero de producto seleccionado
def selectProd(dataframe):
    print("Menu de Productos")
    reduce(lambda none,name: print(f"{name[0]+1}. {name[1]}"),enumerate(dataframe.sheetnames),None)
    return  int(input('Seleccione el Inventario de un Producto: '))

# switch case principal(inicia la operacion seleccionada en el menu de operaciones)
def switch_Main(excel_dataframe,case):
    #en la opcion de inventario inicial se pasa como argumento un libro recien creado,
    #   los demas operaciones reciben un libro seleccionado
    return inventario_inicial(excel_dataframe.create_sheet(input('Ingresa el Nombre del Producto: ')),
                                excel_dataframe) if case==1 else case234(excel_dataframe,case)

# casos 2,3 y 4 de las operaciones 
def case234(excel_dataframe,case):
    return venta(seleccionar_libro(excel_dataframe),excel_dataframe) if  case==2 else  Compra(seleccionar_libro(excel_dataframe),
        excel_dataframe) if case==3 else mostrarInventario(excel_dataframe.worksheets[selectProd(excel_dataframe)-1],leer_excel,[])

# seleccion del producto a operar, devuelve el libro de excel seleccionado(hay un libro por producto)
def seleccionar_libro(excel_dataframe):
    return excel_dataframe.worksheets[selectProd(excel_dataframe)-1]

# lista de las columnas para seleccionar la coluna atraves de un numero N-1
def columnas():
    return ['A','B','C','D','E','F','G','H','I','J']

# muestra las ultimas modificaciones del inventario en la operacion seleccionada 
def ver_ultima_Fila(dataframe,_fila):
    reduce(lambda none,colu:_fila.append(colu[dataframe.max_row-1].value),dataframe.iter_cols(1,dataframe.max_column),None)#leeXcolumna
    print(tabulate([_fila],headers=columnas_nombre(),tablefmt='fancy_grid'))

def columnas_nombre():
    return ['Fecha','Operacion','Entradas','Salidas','Existencia','Unitario','Promedio','Debe','Haber','Saldos']
#------------------------------------------------------------------------------------------------------------------------------------
# Proceso de inventario inicial en inventario 
def inventario_inicial(hoja_trabajo,exceldataframe):
    print("\nInventario inicial")
    
    #encabesados de las columnas
    hoja_trabajo.cell(row=1, column=3, value='Unidades')
    hoja_trabajo.merge_cells('C1:E1')
    hoja_trabajo.cell(row=1, column=6, value='Costo')
    hoja_trabajo.merge_cells('F1:G1')
    hoja_trabajo.cell(row=1, column=8, value='Saldos Finales')
    hoja_trabajo.merge_cells('H1:J1')

    hoja_trabajo.cell(row=2, column=1, value='Fecha')
    hoja_trabajo.cell(row=2, column=2, value='Operacion')
    hoja_trabajo.cell(row=2, column=3, value='Entradas')
    hoja_trabajo.cell(row=2, column=4, value='Salidas')
    hoja_trabajo.cell(row=2, column=5, value='Existencia')
    hoja_trabajo.cell(row=2, column=6, value='Unitario')
    hoja_trabajo.cell(row=2, column=7, value='Promedio')
    hoja_trabajo.cell(row=2, column=8, value='Debe')
    hoja_trabajo.cell(row=2, column=9, value='Haber')
    hoja_trabajo.cell(row=2, column=10, value='Saldos')
    exceldataframe.save("Inventario.xlsx")# guarda los cambios en el excel 
    #fecha y tipo de operacion 
    hoja_trabajo.cell(row=hoja_trabajo.max_row+1, column=1, value=time.strftime('%x'))
    hoja_trabajo.cell(row=hoja_trabajo.max_row, column=2, value='Inv.Inicial')
    
    # ingresa las entradas y el costo unitario 
    reduce(lambda none,columna:hoja_trabajo.cell(row=hoja_trabajo.max_row, column=columna, value=entradas_II(columna)),
            range(3,hoja_trabajo.max_column),None)
    exceldataframe.save("Inventario.xlsx")# guarda los cambios en el excel 
    # realiza las demas operaciones en el inventario
    reduce(lambda none,columna:hoja_trabajo.cell(row=hoja_trabajo.max_row, column=columna, value=existencia_II(columna,hoja_trabajo)),
            range(3,hoja_trabajo.max_column+1),None)
    exceldataframe.save("Inventario.xlsx")# guarda los cambios en el excel 
    ver_ultima_Fila(hoja_trabajo,[])# mostrar los datos agregados
    switch_Main(excel_dataframe(),selectOper())# de buelt al menu principal  

# estas funciones permiten el ingreso de los vaores del inventario inicial y de la compra
def entradas_II(clave):
    return int(input("ingrese las entradas: ")) if clave==3 else unitario_II(clave)

def unitario_II(clave):
    return int(input("ingrese el costo unitario: ")) if clave==6 else 0

# estas funciones devuelven los vaores calculados del inventario inicial
def existencia_II(clave,hoja_trabajo):
    return int(hoja_trabajo[f'C{hoja_trabajo.max_row}'].value) if clave==5 else promedio_II(clave,hoja_trabajo)

def promedio_II(clave,hoja_trabajo):
    return int(hoja_trabajo[f'F{hoja_trabajo.max_row}'].value) if clave==7 else debe_II(clave,hoja_trabajo)

def debe_II(clave,hoja_trabajo):
    return int(hoja_trabajo[f'F{hoja_trabajo.max_row}'].value) * int(hoja_trabajo[f'C{hoja_trabajo.max_row}'].value) if clave==8 else saldos_II(clave,hoja_trabajo)

def saldos_II(clave,hoja_trabajo):
    return int(hoja_trabajo[f'F{hoja_trabajo.max_row}'].value) * int(hoja_trabajo[f'C{hoja_trabajo.max_row}'].value) if clave==10 else mismo_valor(hoja_trabajo,clave)

def mismo_valor(hoja_trabajo,clave):
    return hoja_trabajo[f'{columnas()[clave-1]}{hoja_trabajo.max_row}'].value
#------------------------------------------------------------------------------------------------------------------------------------
# Proceso de compra en inventario 
def Compra(hoja_trabajo,exceldataframe):
    print("\nCompra")
    #fecha y tipo de operacion 
    hoja_trabajo.cell(row=hoja_trabajo.max_row+1, column=1, value=time.strftime('%x'))
    hoja_trabajo.cell(row=hoja_trabajo.max_row, column=2, value='Compra')

    reduce(lambda none,columna:hoja_trabajo.cell(row=hoja_trabajo.max_row, column=columna, value=entradas_II(columna)),
            range(3,hoja_trabajo.max_column+1),None)
    exceldataframe.save("Inventario.xlsx")# guarda los cambios en el excel 
    # realiza las demas operaciones en el inventario
    reduce(lambda none,columna:hoja_trabajo.cell(row=hoja_trabajo.max_row, column=columna, value=existencia_C(columna,hoja_trabajo)),
            range(3,hoja_trabajo.max_column+1),None)
    exceldataframe.save("Inventario.xlsx")# guarda los cambios en el excel 
    reduce(lambda none,columna:hoja_trabajo.cell(row=hoja_trabajo.max_row, column=columna, value=saldos_C(columna,hoja_trabajo)),
            range(3,hoja_trabajo.max_column+1),None)
    exceldataframe.save("Inventario.xlsx")# guarda los cambios en el excel 
    ver_ultima_Fila(hoja_trabajo,[])# mostrar los datos agregados
    switch_Main(excel_dataframe(),selectOper())# de buelt al menu principal

# estas funciones devuelven los vaores calculados de la compra
def existencia_C(clave,hoja_trabajo):
    return int(hoja_trabajo[f'C{hoja_trabajo.max_row}'].value) + int(hoja_trabajo[f'E{hoja_trabajo.max_row-1}'].value) if clave==5 else debe_C(clave,hoja_trabajo)

def debe_C(clave,hoja_trabajo):
    return int(hoja_trabajo[f'F{hoja_trabajo.max_row}'].value) * int(hoja_trabajo[f'C{hoja_trabajo.max_row}'].value) if clave==8 else mismo_valor(hoja_trabajo,clave)

def saldos_C(clave,hoja_trabajo):
    return int(hoja_trabajo[f'H{hoja_trabajo.max_row}'].value) + int(hoja_trabajo[f'J{hoja_trabajo.max_row-1}'].value)  if clave==10 else promedio_C(clave,hoja_trabajo)

def promedio_C(clave,hoja_trabajo):
    return round((int(hoja_trabajo[f'H{hoja_trabajo.max_row}'].value) + int(hoja_trabajo[f'J{hoja_trabajo.max_row-1}'].value)) / int(hoja_trabajo[f'E{hoja_trabajo.max_row}'].value)
                ,2)  if clave==7 else salidas_haber(clave,hoja_trabajo) 

def salidas_haber(clave,hoja_trabajo):
    return 0 if clave==4 | 9 else mismo_valor(hoja_trabajo,clave)

#--------------------------------------------------------------------------------------------------------------------------------------
# Proceso de venta en inventario 
def venta(hoja_trabajo,exceldataframe):
    print("\nVenta")
    #fecha y tipo de operacion 
    hoja_trabajo.cell(row=hoja_trabajo.max_row+1, column=1, value=time.strftime('%x'))
    hoja_trabajo.cell(row=hoja_trabajo.max_row, column=2, value='Venta')

    # asignar valor a las salidas 
    hoja_trabajo.cell(row=hoja_trabajo.max_row, column=4, value=int(input("ingrese las salidas: ")))
    exceldataframe.save("Inventario.xlsx")# guarda los cambios en el excel
    reduce(lambda none,columna:hoja_trabajo.cell(row=hoja_trabajo.max_row, column=columna, value=existencia_V(columna,hoja_trabajo)),
            range(3,hoja_trabajo.max_column+1),None)
    exceldataframe.save("Inventario.xlsx")# guarda los cambios en el excel
    ver_ultima_Fila(hoja_trabajo,[])# mostrar los datos agregados
    switch_Main(excel_dataframe(),selectOper())# de buelt al menu principal}

# estas funciones devuelven los vaores calculados de la compra
def existencia_V(clave,hoja_trabajo):
    return int(hoja_trabajo[f'E{hoja_trabajo.max_row-1}'].value) - int(hoja_trabajo[f'D{hoja_trabajo.max_row}'].value) if clave==5 else unitario_V(clave,hoja_trabajo)

def unitario_V(clave,hoja_trabajo):
    return int(hoja_trabajo[f'G{hoja_trabajo.max_row-1}'].value) if clave==6 or clave==7 else  haber_V(clave,hoja_trabajo)

def haber_V(clave,hoja_trabajo):
    return DnXGm(hoja_trabajo) if clave==9 else saldos_V(clave,hoja_trabajo)

def saldos_V(clave,hoja_trabajo):
    return int(hoja_trabajo[f'J{hoja_trabajo.max_row-1}'].value) - DnXGm(hoja_trabajo)  if clave==10 else cero_MV(clave,hoja_trabajo)

def cero_MV(clave,hoja_trabajo):
    return 0 if clave==8 or clave==3 else hoja_trabajo[f'{columnas()[clave-1]}{hoja_trabajo.max_row}'].value

def DnXGm(hoja_trabajo):
    return int(hoja_trabajo[f'D{hoja_trabajo.max_row}'].value) * int(hoja_trabajo[f'G{hoja_trabajo.max_row-1}'].value)
#---------------------------------------------------------------------------------------------------------------
# Mostrar Inventario de un producto
def mostrarInventario(dataframe,leer_excel,data):
    reduce(lambda none,fila:leer_excel(dataframe,[fila-1,],fila,data),range(2,dataframe.max_row),None)#lee las filas 
    print(tabulate(data,headers=columnas_nombre(),tablefmt='fancy_grid'))
    switch_Main(excel_dataframe(),selectOper())# de buelt al menu principal

def leer_excel(dataframe,_fila,fila,data):
    reduce(lambda none,colu:_fila.append(colu[fila].value),dataframe.iter_cols(1,dataframe.max_column),None)#lee las columnas
    data.append(_fila)
#---------------------------------------------------------------------------------------------------------------
def iniciar_Operaciones(libro):
    inventario_inicial(libro[excel_dataframe().sheetnames[0]], 
                       libro) if libro[excel_dataframe().sheetnames[0]]['C1'].value is None  else switch_Main(excel_dataframe(),selectOper()) 

# Main
iniciar_Operaciones(excel_dataframe())




